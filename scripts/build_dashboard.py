from __future__ import annotations

import json
import argparse
import math
import os
import re
import subprocess
import sys
import tempfile
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse
from urllib.request import Request, urlopen

import pandas as pd
from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
INPUT_DIR = PROJECT_DIR / "input"
OUTPUT_DIR = PROJECT_DIR / "output"
OUTPUT_FILE = OUTPUT_DIR / "packaging_dashboard.html"
COMPARISON_OUTPUT_FILE = OUTPUT_DIR / "comparison_dashboard.html"
DAILY_KPI_FILE = OUTPUT_DIR / "daily_kpi.json"

TARGET_RATIO = 0.75

COLUMN_ALIASES = {
    "date": ["datum", "date", "day", "den"],
    "geosize": ["geosize", "geo size", "velikost", "size"],
    "station": ["stanice baleni", "stanica balenia", "packing station", "station"],
    "packing_group": ["baliaca skupina", "balici skupina", "packing group", "packaging group", "skupina"],
    "doprava": ["doprava", "detail dopravy", "transport detail", "delivery detail"],
    "ab_eliminated": ["ab eliminovane", "ab eliminated", "eliminovane", "eliminated"],
    "total_count": ["celkovy pocet", "total count", "count", "pocet", "pocet jobline", "storejoblines", "store job lines"],
}

GEOSIZE_VALUES = {"SPO", "BPO", "XPO", "XL", "VB"}
STATION_VALUES = {"EXPRESS", "EXPRES", "L40", "MO", "SO01", "SOA1", "SOA0", "XXL"}
BALIKOVKA_MARKERS = ("balikovka", "balikovka_den")


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value).strip().lower())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def is_balikovka_workbook(path: Path) -> bool:
    normalized_name = normalize_text(path.name)
    if any(marker in normalized_name for marker in BALIKOVKA_MARKERS):
        return True

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        return any(
            any(marker in normalize_text(sheet_name) for marker in BALIKOVKA_MARKERS)
            for sheet_name in workbook.sheetnames
        )
    except Exception:
        return False


def require_packaging_workbook(path: Path) -> None:
    if is_balikovka_workbook(path):
        raise ValueError(
            "Tento Excel vyzera ako balikovka workbook. "
            "Packaging dashboard ho z bezpecnostnych dovodov odmieta. "
            "Pouzi Data_pro _balení dashboard_6_2026.xlsx alebo iny packaging input."
        )


def find_excel_file() -> Path:
    excel_files = [
        path
        for path in INPUT_DIR.iterdir()
        if path.is_file()
        and path.suffix.lower() in {".xlsx", ".xls", ".xlsm"}
        and not path.name.startswith("~$")
    ]
    if not excel_files:
        raise FileNotFoundError(
            "V zlozke input nie je ziaden Excel subor. Vloz tam .xlsx, .xls alebo .xlsm subor."
        )

    packaging_files = [path for path in excel_files if not is_balikovka_workbook(path)]
    if not packaging_files:
        raise FileNotFoundError(
            "V zlozke input som nasiel len balikovka workbook. "
            "Packaging dashboard ho nepouzije. Vloz packaging Excel typu Data_pro..."
        )

    def workbook_priority(path: Path) -> tuple[int, float, str]:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            preferred = {"SKLC3", "CZLC4"}.issubset(set(workbook.sheetnames))
        except Exception:
            preferred = False
        return (1 if preferred else 0, path.stat().st_mtime, path.name.lower())

    return sorted(packaging_files, key=workbook_priority, reverse=True)[0]


def resolve_excel_input_path(explicit_path: str | None = None) -> Path:
    if explicit_path:
        path = Path(explicit_path).expanduser()
        if not path.is_absolute():
            path = (PROJECT_DIR / path).resolve()
        if not path.exists():
            raise FileNotFoundError(f"Excel subor neexistuje: {path}")
        require_packaging_workbook(path)
        return path

    env_path = os.environ.get("EXCEL_INPUT_PATH", "").strip()
    if env_path:
        path = Path(env_path).expanduser()
        if not path.is_absolute():
            path = (PROJECT_DIR / path).resolve()
        if not path.exists():
            raise FileNotFoundError(f"Excel subor neexistuje: {path}")
        require_packaging_workbook(path)
        return path

    return find_excel_file()


def copy_excel_for_reading(excel_path: Path) -> Path:
    temp_dir = Path(tempfile.gettempdir())
    copy_path = temp_dir / f"dashboard-{excel_path.stem}-{excel_path.stat().st_mtime_ns}.xlsx"
    try:
        shutil.copy2(excel_path, copy_path)
    except (PermissionError, OSError):
        if os.name != "nt":
            raise
        result = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-Command",
                f"Copy-Item -LiteralPath '{str(excel_path)}' -Destination '{str(copy_path)}' -Force",
            ],
            capture_output=True,
            text=True,
        )
        if result.returncode != 0:
            raise PermissionError(result.stderr.strip() or result.stdout.strip())
    return copy_path


def get_streamlit_excel_source() -> str | None:
    source = os.environ.get("EXCEL_SOURCE_URL", "").strip()
    if not source and is_streamlit_runtime():
        try:
            import streamlit as st

            source = str(st.secrets.get("EXCEL_SOURCE_URL", "")).strip()
        except Exception:
            source = ""
    return source or None


def download_excel_to_temp(source_url: str) -> Path:
    parsed = urlparse(source_url)
    suffix = Path(parsed.path).suffix.lower()
    if suffix not in {".xlsx", ".xls", ".xlsm"}:
        suffix = ".xlsx"

    temp_dir = Path(tempfile.gettempdir())
    copy_path = temp_dir / f"dashboard-remote-{abs(hash(source_url))}{suffix}"
    request = Request(source_url, headers={"User-Agent": "Mozilla/5.0"})
    with urlopen(request, timeout=60) as response:
        data = response.read()
    copy_path.write_bytes(data)
    return copy_path


def build_payload_from_excel(excel_path: Path) -> dict[str, Any]:
    require_packaging_workbook(excel_path)
    try:
        records, metadata = load_records(excel_path)
    except Exception:
        readable_copy = copy_excel_for_reading(excel_path)
        try:
            records, metadata = load_records(readable_copy)
        finally:
            try:
                readable_copy.unlink(missing_ok=True)
            except OSError:
                pass
    return build_payload(records, metadata)


def build_payload_from_source(source_url: str) -> dict[str, Any]:
    readable_copy = download_excel_to_temp(source_url)
    try:
        require_packaging_workbook(readable_copy)
        records, metadata = load_records(readable_copy)
    finally:
        try:
            readable_copy.unlink(missing_ok=True)
        except OSError:
            pass
    payload = build_payload(records, metadata)
    payload["metadata"]["source_url"] = source_url
    return payload


def is_streamlit_runtime() -> bool:
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx
    except Exception:
        return False
    return get_script_run_ctx() is not None


def render_streamlit_dashboard() -> None:
    import streamlit as st
    source_url = get_streamlit_excel_source()

    st.set_page_config(page_title="Balení dashboard", layout="wide")
    if source_url:
        st.info(
            "Streamlit číta dáta priamo z externého Excelu. n8n má len aktualizovať "
            "súbor na stabilnej URL alebo v secrese `EXCEL_SOURCE_URL`."
        )
        try:
            payload = build_payload_from_source(source_url)
        except Exception as error:
            st.warning(
                "Live Excel sa nepodarilo načítať, preto prepínam na lokálny snapshot z docs/."
            )
            st.caption(f"Chyba pri načítaní zdroja: {error}")
        else:
            st.components.v1.html(render_html(payload), height=1600, scrolling=True)
            st.caption(
                f"Zdroj: {payload['metadata'].get('source_file', 'unknown')} | "
                f"URL: {source_url}"
            )
            return

    docs_dir = PROJECT_DIR / "docs"
    parent_file = docs_dir / "index.html"
    child_file = docs_dir / "vyvoj-balenia.html"

    if not parent_file.exists():
        st.error("Nenasiel som docs/index.html. Najprv vytvor nadradenu stranku.")
        return
    if not child_file.exists():
        st.error("Nenasiel som docs/vyvoj-balenia.html. Najprv obnov snapshot detailu.")
        return

    import html as html_lib

    parent_html = parent_file.read_text(encoding="utf-8")
    child_html = child_file.read_text(encoding="utf-8")
    rendered_html = parent_html.replace(
        'src="vyvoj-balenia.html"',
        f'srcdoc="{html_lib.escape(child_html, quote=True)}"',
        1,
    )

    st.info(
        "Používam lokálny snapshot z docs/. Ak nastavíš `EXCEL_SOURCE_URL`, appka "
        "prejde na živé dáta z Excelu."
    )
    st.components.v1.html(rendered_html, height=1600, scrolling=True)
    st.caption("Vývoj balenia je uložený ako stabilný snapshot a dá sa sem pridávať ďalšími kartami.")


def infer_columns(columns: list[Any]) -> dict[str, str | None]:
    normalized_columns = {column: normalize_text(column) for column in columns}
    result: dict[str, str | None] = {}
    for field, aliases in COLUMN_ALIASES.items():
        normalized_aliases = [normalize_text(alias) for alias in aliases]
        exact = next(
            (
                column
                for column, normalized in normalized_columns.items()
                if normalized in normalized_aliases
            ),
            None,
        )
        if exact is not None:
            result[field] = str(exact)
            continue
        partial = next(
            (
                column
                for column, normalized in normalized_columns.items()
                if any(alias in normalized or normalized in alias for alias in normalized_aliases)
            ),
            None,
        )
        result[field] = str(partial) if partial is not None else None
    return result


def safe_number(value: Any, default: float = 0) -> float:
    if pd.isna(value):
        return default
    if isinstance(value, (int, float)):
        return 0 if math.isnan(value) else float(value)
    text = str(value).strip().lower().replace(",", ".")
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def as_dimension(value: Any, fallback: str = "Nezadane") -> str:
    if pd.isna(value):
        return fallback
    text = str(value).strip()
    return text if text else fallback


def as_date(value: Any) -> str:
    if pd.isna(value):
        return "Nezadane"
    text = str(value).strip()
    dayfirst = not bool(re.match(r"^\d{4}-\d{1,2}-\d{1,2}$", text))
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=dayfirst)
    if pd.isna(parsed):
        return as_dimension(value)
    return parsed.strftime("%Y-%m-%d")


def is_numeric_value(value: Any) -> bool:
    if pd.isna(value):
        return False
    try:
        float(value)
        return True
    except (TypeError, ValueError):
        return False


def get_cell_indent(cell: Any) -> int:
    indent = getattr(getattr(cell, "alignment", None), "indent", 0) or 0
    try:
        return int(indent)
    except (TypeError, ValueError):
        return 0


def find_pivot_header_row(worksheet: Any) -> int | None:
    for row_index in range(1, worksheet.max_row + 1):
        value = worksheet.cell(row_index, 1).value
        if "popisky radku" in normalize_text(value):
            return row_index
    return None


def row_has_numeric_values(worksheet: Any, row_index: int, date_columns: list[int]) -> bool:
    return any(is_numeric_value(worksheet.cell(row_index, column_index).value) for column_index in date_columns)


def parse_pivot_sheet(worksheet: Any, sheet_name: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    header_row = find_pivot_header_row(worksheet)
    if header_row is None:
        return [], {}

    date_columns: list[tuple[int, str]] = []
    for column_index in range(2, worksheet.max_column + 1):
        parsed = as_date(worksheet.cell(header_row, column_index).value)
        if parsed != "Nezadane":
            date_columns.append((column_index, parsed))

    records: list[dict[str, Any]] = []
    current_doprava = "Nezadane"
    current_geosize = "Nezadane"
    current_station = "Nezadane"
    current_packing_group = "Nezadane"

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        label = as_dimension(worksheet.cell(row_index, 1).value)
        if label == "Nezadane":
            continue
        normalized_label = normalize_text(label)
        if normalized_label in {"celkovy soucet", "grand total"}:
            continue

        indent = get_cell_indent(worksheet.cell(row_index, 1))
        row_has_values = row_has_numeric_values(worksheet, row_index, [column for column, _ in date_columns])
        upper_label = label.upper()

        if indent == 0 and not row_has_values and upper_label not in GEOSIZE_VALUES and upper_label not in STATION_VALUES:
            current_doprava = label
            current_geosize = "Nezadane"
            current_station = "Nezadane"
            current_packing_group = "Nezadane"
            continue
        if upper_label in GEOSIZE_VALUES:
            current_geosize = upper_label
            current_station = "Nezadane"
            current_packing_group = "Nezadane"
            continue
        if upper_label in STATION_VALUES:
            current_station = "Expres" if upper_label in {"EXPRESS", "EXPRES"} else upper_label
            current_packing_group = "Nezadane"
            continue
        if indent >= 3 and not row_has_values:
            current_packing_group = label
            continue
        if not row_has_values:
            continue

        is_ab = "ab eliminovane" in normalized_label
        doprava = current_doprava

        for column_index, date_value in date_columns:
            count = safe_number(worksheet.cell(row_index, column_index).value)
            if count <= 0:
                continue
            records.append(
                {
                    "date": date_value,
                    "geosize": current_geosize,
                    "station": current_station,
                    "packing_group": current_packing_group if indent >= 4 else label,
                    "doprava": doprava,
                    "ab_eliminated": count if is_ab else 0,
                    "total_count": count,
                    "sheet": sheet_name,
                }
            )

    detected = {
        "layout": "pivot",
        "date": "Popisky sloupcu",
        "geosize": "Popisky radku uroven 1",
        "station": "Popisky radku uroven 2",
        "packing_group": "Popisky radku uroven 3",
        "doprava": "Doprava",
        "ab_eliminated": "AB Eliminovane etikety",
        "total_count": "Pocet jobline",
    }
    return records, detected


def load_records(excel_path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    sheets = pd.read_excel(excel_path, sheet_name=None)
    workbook = load_workbook(excel_path, data_only=True)
    records: list[dict[str, Any]] = []
    sheet_info: list[dict[str, Any]] = []
    detected_columns: dict[str, Any] = {}

    for sheet_name, frame in sheets.items():
        if frame.empty:
            continue

        inferred = infer_columns(list(frame.columns))
        standard_layout = bool(inferred.get("ab_eliminated") or inferred.get("total_count"))
        if not standard_layout:
            pivot_records, pivot_detected = parse_pivot_sheet(workbook[sheet_name], sheet_name)
            if pivot_records:
                records.extend(pivot_records)
                raw_sheet = workbook[sheet_name]
                sheet_info.append(
                    {
                        "sheet": sheet_name,
                        "rows": raw_sheet.max_row,
                        "columns": raw_sheet.max_column,
                        "detected": pivot_detected,
                    }
                )
                detected_columns = detected_columns or pivot_detected
                continue

        sheet_info.append(
            {
                "sheet": sheet_name,
                "rows": len(frame),
                "columns": len(frame.columns),
                "detected": inferred,
            }
        )
        detected_columns = detected_columns or inferred

        for _, row in frame.iterrows():
            total_column = inferred.get("total_count")
            eliminated_column = inferred.get("ab_eliminated")
            total = safe_number(row.get(total_column), 1) if total_column else 1
            eliminated = safe_number(row.get(eliminated_column), 0) if eliminated_column else 0
            if total <= 0 and eliminated <= 0:
                continue
            records.append(
                {
                    "date": as_date(row.get(inferred["date"])) if inferred.get("date") else "Nezadane",
                    "geosize": as_dimension(row.get(inferred["geosize"])) if inferred.get("geosize") else "Nezadane",
                    "station": as_dimension(row.get(inferred["station"])) if inferred.get("station") else "Nezadane",
                    "packing_group": as_dimension(row.get(inferred["packing_group"])) if inferred.get("packing_group") else "Nezadane",
                    "doprava": as_dimension(row.get(inferred["doprava"])) if inferred.get("doprava") else "Nezadane",
                    "ab_eliminated": eliminated,
                    "total_count": total,
                    "sheet": sheet_name,
                }
            )

    metadata = {
        "source_file": excel_path.name,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "target_ratio": TARGET_RATIO,
        "sheet_info": sheet_info,
        "detected_columns": detected_columns,
    }
    return records, metadata


def build_payload(records: list[dict[str, Any]], metadata: dict[str, Any]) -> dict[str, Any]:
    return {"metadata": metadata, "records": records}


def is_special_elimination_group(row: dict[str, Any]) -> bool:
    return normalize_text(row.get("packing_group")) == "spo pob dist nebalit standard"


def elimination_count(row: dict[str, Any]) -> float:
    ab = safe_number(row.get("ab_eliminated"), 0)
    special_group = safe_number(row.get("total_count"), 0) if is_special_elimination_group(row) else 0
    return ab + special_group


def build_comparison_summary(records: list[dict[str, Any]]) -> dict[str, Any]:
    sheet_names = ["SKLC3", "CZLC4"]
    summaries: list[dict[str, Any]] = []
    combined_total = sum(safe_number(row.get("total_count"), 0) for row in records)
    for sheet_name in sheet_names:
        rows = [row for row in records if normalize_text(row.get("sheet")) == normalize_text(sheet_name)]
        total = sum(safe_number(row.get("total_count"), 0) for row in rows)
        eliminated = sum(elimination_count(row) for row in rows)
        summaries.append(
            {
                "sheet": sheet_name,
                "total": total,
                "share": (total / combined_total) if combined_total else 0,
                "rows": len(rows),
                "eliminated": eliminated,
                "elimination_share": (eliminated / total) if total else 0,
            }
        )
    summaries.sort(key=lambda item: item["total"], reverse=True)
    return {
        "combined_total": combined_total,
        "summaries": summaries,
    }


def build_daily_kpi_summary(records: list[dict[str, Any]]) -> dict[str, Any]:
    available_days = sorted(
        {
            str(row.get("date"))
            for row in records
            if isinstance(row.get("date"), str)
            and re.fullmatch(r"\d{4}-\d{2}-\d{2}", str(row.get("date")))
        }
    )
    fallback_day = (datetime.now().date() - pd.Timedelta(days=1)).isoformat()
    target_day = available_days[-1] if available_days else fallback_day
    daily_records = [row for row in records if str(row.get("date")) == target_day]
    sheet_rows: list[dict[str, Any]] = []
    mail_lines: list[str] = []
    for sheet_name in ["SKLC3", "CZLC4"]:
        filtered = [
            row
            for row in daily_records
            if normalize_text(row.get("sheet")) == normalize_text(sheet_name)
            and normalize_text(row.get("geosize")) == normalize_text("SPO")
            and normalize_text(row.get("doprava")) == normalize_text("Alzabox")
        ]
        total_count = sum(safe_number(row.get("total_count"), 0) for row in filtered)
        eliminated_count = sum(elimination_count(row) for row in filtered)
        sheet_rows.append(
            {
                "sheet": sheet_name,
                "total_count": total_count,
                "eliminated_count": eliminated_count,
                "ratio": (eliminated_count / total_count) if total_count else None,
            }
        )
        ratio = sheet_rows[-1]["ratio"]
        ratio_text = "bez dat" if ratio is None else f"{ratio * 100:.1f}".replace(".", ",") + " %"
        mail_lines.append(
            f"{sheet_name}: {ratio_text} "
            f"({format_int_text(eliminated_count)} / {format_int_text(total_count)}) "
            f"eliminace z geosize = SPO, doprava = alzabox"
        )
    return {
        "target_day": target_day,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "sheet_rows": sheet_rows,
        "mail_lines": mail_lines,
    }


def format_int_text(value: Any) -> str:
    return f"{int(round(safe_number(value, 0))):,}".replace(",", " ")


def format_pct_text(value: Any) -> str:
    return f"{safe_number(value, 0) * 100:.1f}".replace(".", ",") + " %"


def escape_html(value: Any) -> str:
    return (
        str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#039;")
    )


def to_json_for_html(payload: dict[str, Any]) -> str:
    return json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")


def render_html(payload: dict[str, Any]) -> str:
    data_json = to_json_for_html(payload)
    html = """<!doctype html>
<html lang="sk">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Vyvoj balenia dashboard</title>
  <style>
    :root {
      --bg: #f5f7fa;
      --panel: #ffffff;
      --ink: #15202b;
      --muted: #5f6b7a;
      --line: #d8dee8;
      --blue: #2563eb;
      --green: #0f9f6e;
      --amber: #d97706;
      --red: #dc2626;
      --shadow: 0 12px 30px rgba(23, 37, 84, .08);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font-family: "Segoe UI", Arial, sans-serif;
    }
    header {
      background: #111827;
      color: white;
      padding: 24px clamp(16px, 3vw, 36px);
    }
    header h1 {
      margin: 0 0 8px;
      font-size: clamp(24px, 3vw, 36px);
      letter-spacing: 0;
    }
    header p { margin: 0; color: #cbd5e1; }
    main {
      width: min(1440px, 100%);
      margin: 0 auto;
      padding: 24px clamp(12px, 2.5vw, 32px) 36px;
    }
    .meta {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      color: var(--muted);
      font-size: 13px;
      margin-bottom: 16px;
    }
    .pill {
      background: #eaf0f8;
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 6px 10px;
    }
    .dashboards {
      display: grid;
      gap: 22px;
    }
    .dashboard {
      display: grid;
      gap: 14px;
      padding: 18px;
      background: #eef2f7;
      border: 1px solid var(--line);
      border-radius: 8px;
    }
    .dashboard-head {
      display: flex;
      align-items: flex-start;
      justify-content: space-between;
      gap: 16px;
    }
    .dashboard-head h2 {
      margin: 0 0 4px;
      font-size: 20px;
      letter-spacing: 0;
    }
    .dashboard-head p {
      margin: 0;
      color: var(--muted);
      font-size: 13px;
    }
    .filters {
      display: grid;
      grid-template-columns: repeat(5, minmax(160px, 1fr));
      gap: 12px;
    }
    label {
      display: grid;
      gap: 6px;
      color: var(--muted);
      font-size: 12px;
      font-weight: 600;
      text-transform: uppercase;
    }
    select {
      min-height: 38px;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 7px 10px;
      background: white;
      color: var(--ink);
      font-size: 14px;
      width: 100%;
    }
    .cards {
      display: grid;
      grid-template-columns: repeat(4, minmax(170px, 1fr));
      gap: 14px;
    }
    .card, .panel {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
    }
    .card { padding: 16px; }
    .card span {
      color: var(--muted);
      font-size: 12px;
      font-weight: 700;
      text-transform: uppercase;
    }
    .card strong {
      display: block;
      margin-top: 8px;
      font-size: clamp(24px, 2.8vw, 34px);
    }
    .card small { color: var(--muted); }
    .goalbar {
      height: 9px;
      background: #e5e7eb;
      border-radius: 999px;
      overflow: hidden;
      margin-top: 12px;
    }
    .goalbar div {
      height: 100%;
      width: 0;
      background: var(--green);
      transition: width .2s ease;
    }
    .grid {
      display: grid;
      grid-template-columns: minmax(0, 1.35fr) minmax(340px, .65fr);
      gap: 16px;
      align-items: start;
    }
    .panel {
      padding: 16px;
      min-width: 0;
    }
    .panel h2 {
      margin: 0 0 14px;
      font-size: 18px;
      letter-spacing: 0;
    }
    .chart-wrap {
      width: 100%;
      height: 360px;
    }
    svg { width: 100%; height: 100%; display: block; }
    .table-wrap { overflow: auto; }
    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
    }
    th, td {
      padding: 9px 8px;
      border-bottom: 1px solid var(--line);
      text-align: right;
      white-space: nowrap;
    }
    th:first-child, td:first-child { text-align: left; }
    th {
      color: var(--muted);
      font-size: 12px;
      text-transform: uppercase;
    }
    .ratio.good { color: var(--green); font-weight: 700; }
    .ratio.mid { color: var(--amber); font-weight: 700; }
    .ratio.bad { color: var(--red); font-weight: 700; }
    .empty {
      padding: 28px;
      text-align: center;
      color: var(--muted);
      border: 1px dashed var(--line);
      border-radius: 8px;
      background: #fbfcfe;
    }
    @media (max-width: 1020px) {
      .filters, .cards, .grid { grid-template-columns: 1fr 1fr; }
      .grid .panel:first-child { grid-column: 1 / -1; }
      .dashboard-head { flex-direction: column; }
    }
    @media (max-width: 680px) {
      .filters, .cards, .grid { grid-template-columns: 1fr; }
      header { padding-top: 18px; padding-bottom: 18px; }
      .chart-wrap { height: 300px; }
    }
  </style>
</head>
<body>
  <header>
    <h1>Vyvoj balenia</h1>
    <p>AB eliminated / total count, target 75 %</p>
  </header>
  <main>
    <div class="meta" id="meta"></div>
    <div class="dashboards" id="dashboards"></div>
  </main>
  <script id="dashboard-data" type="application/json">__DATA_JSON__</script>
  <script>
    const payload = JSON.parse(document.getElementById('dashboard-data').textContent);
    const records = payload.records || [];
    const targetRatio = payload.metadata?.target_ratio ?? 0.75;
    const balikovkaTarget = 0.10;
    const sheetOrder = (payload.metadata?.sheet_info || []).map(info => info.sheet).filter(Boolean);
    const dashboardSheets = sheetOrder.length
      ? sheetOrder
      : [...new Set(records.map(row => row.sheet).filter(Boolean))];
    const balikovkaTransports = {
      DEFAULT: [
        'DHL',
        'DPD',
        'B2B',
        'AlzaExpres',
        'ExpressOne',
        'FoxPost',
        'Gebruder Weiss',
        'GO!',
        'Maďarská pošta',
        'MyFlexBox',
        'Post AT',
        'PostAT',
        'Pošta',
        'Pošta Slovensko',
        'PPL',
        'SPS',
        'TopTrans',
        'WeDo',
        'Zásilkovna',
      ],
      SKLC3: [
        'DHL',
        'DPD',
        'B2B',
        'AlzaExpres',
        'ExpressOne',
        'FoxPost',
        'Gebruder Weiss',
        'GO!',
        'Maďarská pošta',
        'MyFlexBox',
        'Post AT',
        'Pošta',
        'Pošta Slovensko',
        'PPL',
        'SPS',
        'TopTrans',
        'WeDo',
        'Zásilkovna',
      ],
      CZLC4: [
        'DHL',
        'DPD',
        'B2B',
        'AlzaExpres',
        'ExpressOne',
        'FoxPost',
        'Gebruder Weiss',
        'GO!',
        'Kurýr',
        'Maďarská pošta',
        'MyFlexBox',
        'Post AT',
        'Pošta',
        'Pošta Slovensko',
        'PPL',
        'SPS',
        'TopTrans',
        'WeDo',
        'Zásilkovna',
      ],
    };

    const labels = {
      date: 'Date',
      geosize: 'Geosize',
      station: 'Packing station',
      packing_group: 'Packing group',
      doprava: 'Transport detail',
    };

    const fmtInt = new Intl.NumberFormat('sk-SK', { maximumFractionDigits: 0 });
    const fmtPct = new Intl.NumberFormat('sk-SK', {
      minimumFractionDigits: 1,
      maximumFractionDigits: 1,
    });

    function formatInt(value) {
      return fmtInt.format(Math.round(Number(value) || 0));
    }

    function formatPct(value) {
      return fmtPct.format((Number(value) || 0) * 100) + ' %';
    }

    function ratioClass(value) {
      if (value >= targetRatio) return 'good';
      if (value >= 0.65) return 'mid';
      return 'bad';
    }

    function balikovkaClass(value) {
      if (value <= balikovkaTarget) return 'good';
      if (value <= 0.12) return 'mid';
      return 'bad';
    }

    function slugify(value) {
      return String(value ?? '')
        .toLowerCase()
        .normalize('NFKD')
        .replace(/[\u0300-\u036f]/g, '')
        .replace(/[^a-z0-9]+/g, '-')
        .replace(/^-+|-+$/g, '');
    }

    function normalizeText(value) {
      return String(value ?? '')
        .normalize('NFKD')
        .replace(/[\u0300-\u036f]/g, '')
        .toLowerCase()
        .trim();
    }

    function escapeHtml(value) {
      return String(value).replace(/[&<>"']/g, char => ({
        '&': '&amp;', '<': '&lt;', '>': '&gt;', '"': '&quot;', "'": '&#039;'
      }[char]));
    }

    function normalizeLabel(value) {
      return String(value ?? 'Nezadane');
    }

    function parseDateKey(value) {
      if (value instanceof Date && !Number.isNaN(value.getTime())) {
        return value;
      }
      const text = normalizeLabel(value);
      if (!text || text === 'Nezadane') return null;
      const parsed = new Date(`${text}T00:00:00`);
      return Number.isNaN(parsed.getTime()) ? null : parsed;
    }

    function formatTrendDate(value) {
      const parsed = parseDateKey(value);
      if (!parsed) return normalizeLabel(value);
      const day = String(parsed.getDate()).padStart(2, '0');
      const month = String(parsed.getMonth() + 1).padStart(2, '0');
      const year = parsed.getFullYear();
      return `${day}.${month}.${year}`;
    }

    function displayLabel(field, value) {
      const normalized = normalizeLabel(value);
      if (field === 'doprava' && normalized === 'Nezadane') return 'Pobocka';
      return normalized;
    }

    function isSpecialEliminationGroup(row) {
      return normalizeLabel(row.packing_group).toLowerCase() === 'spo pob, dist nebalit standard';
    }

    function eliminationCount(row) {
      const ab = Number(row.ab_eliminated) || 0;
      const specialGroup = isSpecialEliminationGroup(row) ? (Number(row.total_count) || 0) : 0;
      return ab + specialGroup;
    }

    function balikovkaTransportSet(sheetName) {
      const transports = balikovkaTransports[sheetName] || balikovkaTransports.DEFAULT || [];
      return new Set(transports.map(normalizeText));
    }

    function balikovkaCount(rows, sheetName) {
      const allowed = balikovkaTransportSet(sheetName);
      if (!allowed.size) return 0;
      return rows.reduce((sum, row) => {
        const transport = normalizeText(row.doprava);
        return sum + (allowed.has(transport) ? (Number(row.total_count) || 0) : 0);
      }, 0);
    }

    function aggregate(rows) {
      let eliminated = 0;
      let total = 0;
      for (const row of rows) {
        eliminated += eliminationCount(row);
        total += Number(row.total_count) || 0;
      }
      return {
        ab: eliminated,
        total,
        ratio: total ? eliminated / total : 0,
      };
    }

    function allValues(rows, field) {
      return ['all', ...new Set(rows.map(row => normalizeLabel(row[field])))].sort((a, b) => a.localeCompare(b, 'sk'));
    }

    function seriesByDate(rows) {
      const map = new Map();
      for (const row of rows) {
        const key = normalizeLabel(row.date);
        const item = map.get(key) || { key, ab: 0, total: 0 };
        item.ab += eliminationCount(row);
        item.total += Number(row.total_count) || 0;
        map.set(key, item);
      }
      return [...map.values()].sort((a, b) => {
        const aDate = parseDateKey(a.key);
        const bDate = parseDateKey(b.key);
        if (aDate && bDate) return aDate - bDate;
        if (aDate) return -1;
        if (bDate) return 1;
        return a.key.localeCompare(b.key, 'sk');
      }).map(item => ({
        ...item,
        ratio: item.total ? item.ab / item.total : 0,
      }));
    }

    function renderTrend(rows, targetId) {
      const data = seriesByDate(rows);
      const holder = document.getElementById(targetId);
      if (!data.length) {
        holder.innerHTML = '<div class="empty">No data for trend.</div>';
        return;
      }

      const width = 920;
      const height = 340;
      const left = 54;
      const right = 22;
      const top = 18;
      const bottom = 48;
      const innerW = width - left - right;
      const innerH = height - top - bottom;
      const maxY = Math.max(targetRatio, ...data.map(d => d.ratio), 0.8);
      const x = index => left + (data.length === 1 ? innerW / 2 : index * innerW / (data.length - 1));
      const y = value => top + innerH - (value / maxY) * innerH;
      const points = data.map((d, i) => `${x(i)},${y(d.ratio)}`).join(' ');
      const targetY = y(targetRatio);
      const labelsLine = data.map((d, i) => {
        const show = data.length <= 12 || i === 0 || i === data.length - 1 || i % Math.ceil(data.length / 8) === 0;
        return show ? `<text x="${x(i)}" y="${height - 16}" text-anchor="middle" font-size="11" fill="#5f6b7a">${escapeHtml(formatTrendDate(d.key))}</text>` : '';
      }).join('');

      holder.innerHTML = `
        <svg viewBox="0 0 ${width} ${height}" role="img" aria-label="Trend AB eliminated">
          <line x1="${left}" y1="${top}" x2="${left}" y2="${height - bottom}" stroke="#d8dee8"/>
          <line x1="${left}" y1="${height - bottom}" x2="${width - right}" y2="${height - bottom}" stroke="#d8dee8"/>
          <line x1="${left}" y1="${targetY}" x2="${width - right}" y2="${targetY}" stroke="#0f9f6e" stroke-dasharray="6 6"/>
          <text x="${width - right}" y="${targetY - 8}" text-anchor="end" font-size="12" fill="#0f9f6e">target 75 %</text>
          <polyline fill="none" stroke="#2563eb" stroke-width="3" points="${points}"/>
          ${data.map((d, i) => `<circle cx="${x(i)}" cy="${y(d.ratio)}" r="4" fill="#2563eb"><title>${escapeHtml(d.key)}: ${formatPct(d.ratio)}</title></circle>`).join('')}
          ${labelsLine}
        </svg>
      `;
    }

    function renderTable(id, rows, key) {
      const map = new Map();
      let totalVolume = 0;
      for (const row of rows) {
        totalVolume += Number(row.total_count) || 0;
        const name = normalizeLabel(row[key]);
        const item = map.get(name) || { name, ab: 0, total: 0 };
        item.ab += eliminationCount(row);
        item.total += Number(row.total_count) || 0;
        map.set(name, item);
      }
      const data = [...map.values()].map(item => ({
        ...item,
        share: totalVolume ? item.total / totalVolume : 0,
        elimRatio: item.total ? item.ab / item.total : 0,
      })).sort((a, b) => b.total - a.total);
      const table = document.getElementById(id);
      if (!data.length) {
        table.innerHTML = '<tbody><tr><td class="empty">No data for the selected filter.</td></tr></tbody>';
        return;
      }
      table.innerHTML = `
        <thead><tr><th>${labels[key]}</th><th>SJLs</th><th>Share</th><th>Elimination</th></tr></thead>
        <tbody>
          ${data.map(row => `
            <tr>
              <td>${escapeHtml(displayLabel(key, row.name))}</td>
              <td>${formatInt(row.total)}</td>
              <td>${formatPct(row.share)}</td>
              <td class="ratio ${ratioClass(row.elimRatio)}">${formatPct(row.elimRatio)}</td>
            </tr>
          `).join('')}
        </tbody>
      `;
    }

    function renderVolumeTable(id, rows, key) {
      renderTable(id, rows, key);
    }

    function renderDashboard(sheetName, index) {
      const info = (payload.metadata?.sheet_info || []).find(item => item.sheet === sheetName) || {};
      const sheetId = slugify(sheetName) || `sheet-${index + 1}`;
      const sheetRows = records.filter(row => normalizeLabel(row.sheet) === normalizeLabel(sheetName));
      const prefix = `${sheetId}_`;
      const root = document.createElement('section');
      root.className = 'dashboard';
      root.id = `dashboard_${sheetId}`;
      root.innerHTML = `
        <div class="dashboard-head">
          <div>
            <h2>Dashboard ${index + 1} - ${escapeHtml(sheetName)}</h2>
            <p>Same conditions as the first dashboard, rendered from the same workbook.</p>
          </div>
          <div class="meta" id="${prefix}meta"></div>
        </div>
        <section class="filters" id="${prefix}filters"></section>
        <section class="cards">
          <div class="card">
            <span>Elimination share</span>
            <strong id="${prefix}kpiRatio">0 %</strong>
            <small>Target: 75 %</small>
            <div class="goalbar"><div id="${prefix}goalFill"></div></div>
          </div>
          <div class="card">
            <span>Balikovka</span>
            <strong id="${prefix}kpiBalikovka">0 %</strong>
            <small id="${prefix}kpiBalikovkaCount">0 SJLs</small>
            <small>Target: &lt; 10 % of SJLs</small>
          </div>
          <div class="card"><span>Eliminated</span><strong id="${prefix}kpiEliminated">0</strong><small>StoreJobLines</small></div>
          <div class="card"><span>Selected total</span><strong id="${prefix}kpiSelectedTotal">0</strong><small>StoreJobLines</small></div>
          <div class="card"><span>Dashboard total</span><strong id="${prefix}kpiTotal">0</strong><small>StoreJobLines</small></div>
          <div class="card"><span>Gap to target</span><strong id="${prefix}kpiGap">0 b.</strong><small>percentage points</small></div>
        </section>
        <section class="grid">
          <div class="panel">
            <h2>Trend by day</h2>
            <div class="chart-wrap" id="${prefix}trendChart"></div>
          </div>
          <div class="panel">
            <h2>Geosize</h2>
            <div class="table-wrap"><table id="${prefix}geosizeTable"></table></div>
          </div>
          <div class="panel">
            <h2>Packing station</h2>
            <div class="table-wrap"><table id="${prefix}stationTable"></table></div>
          </div>
          <div class="panel">
            <h2>Packing groups</h2>
            <div class="table-wrap"><table id="${prefix}groupTable"></table></div>
          </div>
          <div class="panel">
            <h2>Transport detail</h2>
            <div class="table-wrap"><table id="${prefix}detailTable"></table></div>
          </div>
        </section>
      `;
      document.getElementById('dashboards').appendChild(root);

      const state = {
        date: 'all',
        geosize: 'all',
        station: 'all',
        packing_group: 'all',
        doprava: 'all',
      };

      function filteredRows() {
        return sheetRows.filter(row => {
          return Object.entries(state).every(([field, selected]) => {
            if (selected === 'all') return true;
            return normalizeLabel(row[field]) === selected;
          });
        });
      }

      function renderMeta(rows) {
        const meta = payload.metadata || {};
        const totalRows = sheetRows.length;
        const validDates = rows
          .map(row => parseDateKey(row.date))
          .filter(Boolean)
          .sort((a, b) => a - b);
        const dateRange = validDates.length
          ? `${formatTrendDate(validDates[0])} - ${formatTrendDate(validDates[validDates.length - 1])}`
          : 'bez dátumov';
        document.getElementById(`${prefix}meta`).innerHTML = [
          `Source: ${meta.source_file || 'unknown'}`,
          `Sheet: ${sheetName}`,
          `Generated: ${meta.generated_at || ''}`,
          `Date range: ${dateRange}`,
          `Rows: ${formatInt(rows.length)} / ${formatInt(totalRows)}`,
          `Detected: ${info.detected?.doprava ? 'AB + Total + transport' : (info.detected?.ab_eliminated ? 'AB + Total' : 'pivot/flat')}`,
        ].map(text => `<span class="pill">${text}</span>`).join('');
      }

      function renderFilters() {
        const holder = document.getElementById(`${prefix}filters`);
        holder.innerHTML = Object.keys(labels).map(field => `
          <label>${labels[field]}<select id="${prefix}filter_${field}"></select></label>
        `).join('');
        for (const field of Object.keys(state)) {
          const select = document.getElementById(`${prefix}filter_${field}`);
          if (!select) continue;
          select.innerHTML = allValues(sheetRows, field).map(value => {
            const label = value === 'all' ? 'All' : displayLabel(field, value);
            return `<option value="${escapeHtml(value)}">${escapeHtml(label)}</option>`;
          }).join('');
          select.value = state[field];
          select.onchange = event => {
            state[field] = event.target.value;
            render();
          };
        }
      }

      function renderKpis(rows) {
        const summary = aggregate(rows);
        const dashboardSummary = aggregate(sheetRows);
        const balikovka = balikovkaCount(rows, sheetName);
        const balikovkaRatio = summary.total ? balikovka / summary.total : 0;
        document.getElementById(`${prefix}kpiRatio`).textContent = formatPct(summary.ratio);
        document.getElementById(`${prefix}kpiRatio`).className = 'ratio ' + ratioClass(summary.ratio);
        document.getElementById(`${prefix}kpiBalikovka`).textContent = formatPct(balikovkaRatio);
        document.getElementById(`${prefix}kpiBalikovka`).className = 'ratio ' + balikovkaClass(balikovkaRatio);
        document.getElementById(`${prefix}kpiBalikovkaCount`).textContent = `${formatInt(balikovka)} SJLs`;
        document.getElementById(`${prefix}kpiEliminated`).textContent = formatInt(summary.ab);
        document.getElementById(`${prefix}kpiSelectedTotal`).textContent = formatInt(summary.total);
        document.getElementById(`${prefix}kpiTotal`).textContent = formatInt(dashboardSummary.total);
        document.getElementById(`${prefix}kpiGap`).textContent = ((summary.ratio - targetRatio) * 100).toLocaleString('sk-SK', {
          minimumFractionDigits: 1,
          maximumFractionDigits: 1,
        }) + ' b.';
        document.getElementById(`${prefix}goalFill`).style.width = Math.min(100, (summary.ratio / targetRatio) * 100) + '%';
      }

      function render() {
        const rows = filteredRows();
        renderMeta(rows);
        renderKpis(rows);
        renderFilters();
        renderTrend(rows, `${prefix}trendChart`);
        renderTable(`${prefix}geosizeTable`, rows, 'geosize');
        renderTable(`${prefix}stationTable`, rows, 'station');
        renderTable(`${prefix}groupTable`, rows, 'packing_group');
        renderVolumeTable(`${prefix}detailTable`, rows, 'doprava');
      }

      render();
    }

    document.getElementById('meta').innerHTML = [
      `Source: ${payload.metadata?.source_file || 'unknown'}`,
      `Generated: ${payload.metadata?.generated_at || ''}`,
      `Dashboards: ${dashboardSheets.length || 0}`,
    ].map(text => `<span class="pill">${text}</span>`).join('');

    const dashboardsHolder = document.getElementById('dashboards');
    if (!dashboardSheets.length) {
      dashboardsHolder.innerHTML = '<div class="empty">No records found in the input workbook.</div>';
    } else {
      dashboardSheets.forEach((sheetName, index) => renderDashboard(sheetName, index));
    }
  </script>
</body>
</html>
"""
    return html.replace("__DATA_JSON__", data_json)


def save_dashboard(payload: dict[str, Any]) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_FILE.write_text(render_html(payload), encoding="utf-8")


def render_comparison_html(payload: dict[str, Any]) -> str:
    summary = build_comparison_summary(payload["records"])
    combined_total = summary["combined_total"]
    rows_html = []
    for item in summary["summaries"]:
        rows_html.append(
            f"""
            <article class="card">
              <span>{item['sheet']}</span>
              <strong>{format_int_text(item['total'])} SJLs</strong>
              <small>Podiel z celku: {format_pct_text(item['share'])}</small>
              <small>Riadkov: {format_int_text(item['rows'])}</small>
            </article>
            """
        )

    table_rows = []
    for item in summary["summaries"]:
        table_rows.append(
            f"""
            <tr>
              <td>{escape_html(item['sheet'])}</td>
              <td>{format_int_text(item['total'])}</td>
              <td>{format_pct_text(item['share'])}</td>
              <td>{format_int_text(item['rows'])}</td>
              <td>{format_pct_text(item['elimination_share'])}</td>
            </tr>
            """
        )

    bar_rows = []
    for item in summary["summaries"]:
        bar_rows.append(
            f"""
            <div class="share-row">
              <div class="share-head">
                <strong>{escape_html(item['sheet'])}</strong>
                <span>{format_int_text(item['total'])} SJLs · {format_pct_text(item['share'])}</span>
              </div>
              <div class="share-bar"><div style="width: {min(100, item['share'] * 100):.2f}%"></div></div>
            </div>
            """
        )

    html = f"""<!doctype html>
<html lang="sk">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Porovnanie SKLC3 vs CZLC4</title>
  <style>
    :root {{
      --bg: #f5f7fa;
      --panel: #ffffff;
      --ink: #15202b;
      --muted: #5f6b7a;
      --line: #d8dee8;
      --blue: #2563eb;
      --green: #0f9f6e;
      --shadow: 0 12px 30px rgba(23, 37, 84, .08);
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font-family: "Segoe UI", Arial, sans-serif;
    }}
    header {{
      background: linear-gradient(135deg, #111827, #1f2937);
      color: white;
      padding: 28px clamp(16px, 3vw, 40px);
    }}
    header h1 {{
      margin: 0 0 6px;
      font-size: clamp(26px, 3vw, 40px);
    }}
    header p {{ margin: 0; color: #cbd5e1; }}
    main {{
      width: min(1120px, 100%);
      margin: 0 auto;
      padding: 24px clamp(12px, 2.5vw, 32px) 36px;
    }}
    .meta {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      color: var(--muted);
      font-size: 13px;
      margin-bottom: 16px;
    }}
    .pill {{
      background: #eaf0f8;
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 6px 10px;
    }}
    .cards {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 14px;
      margin-bottom: 18px;
    }}
    .card, .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 16px;
      box-shadow: var(--shadow);
    }}
    .card {{
      padding: 18px 20px;
      display: grid;
      gap: 6px;
    }}
    .card span {{
      font-size: 12px;
      font-weight: 700;
      color: var(--muted);
      text-transform: uppercase;
      letter-spacing: .04em;
    }}
    .card strong {{
      font-size: clamp(28px, 4vw, 48px);
      line-height: 1;
    }}
    .card small {{
      color: var(--muted);
      font-size: 13px;
    }}
    .panel {{
      padding: 18px 20px 22px;
      margin-top: 16px;
    }}
    .panel h2 {{
      margin: 0 0 16px;
      font-size: 20px;
    }}
    .share-row + .share-row {{ margin-top: 16px; }}
    .share-head {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 8px;
      color: var(--ink);
    }}
    .share-head span {{ color: var(--muted); }}
    .share-bar {{
      height: 16px;
      background: #e7edf5;
      border-radius: 999px;
      overflow: hidden;
    }}
    .share-bar div {{
      height: 100%;
      border-radius: 999px;
      background: linear-gradient(90deg, var(--blue), #7c3aed);
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 14px;
    }}
    th, td {{
      text-align: left;
      padding: 12px 0;
      border-bottom: 1px solid var(--line);
    }}
    th {{
      color: var(--muted);
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: .04em;
    }}
    .ratio.good {{ color: var(--green); font-weight: 700; }}
    .ratio.mid {{ color: #d97706; font-weight: 700; }}
    .ratio.bad {{ color: #dc2626; font-weight: 700; }}
    @media (max-width: 860px) {{
      .cards {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>Porovnanie SKLC3 vs CZLC4</h1>
    <p>Objem SJLs a podiel každej záložky z celkového objemu.</p>
  </header>
  <main>
    <div class="meta">
      <span class="pill">Source: {escape_html(payload["metadata"].get("source_file", "unknown"))}</span>
      <span class="pill">Generated: {escape_html(payload["metadata"].get("generated_at", ""))}</span>
      <span class="pill">Combined total: {format_int_text(combined_total)} SJLs</span>
      <span class="pill">Sheets: 2</span>
    </div>
    <section class="cards">
      {''.join(rows_html)}
    </section>
    <section class="panel">
      <h2>Podiel na celku</h2>
      {''.join(bar_rows)}
    </section>
    <section class="panel">
      <h2>Prehľad</h2>
      <table>
        <thead>
          <tr>
            <th>Sheet</th>
            <th>Volume</th>
            <th>Share</th>
            <th>Rows</th>
            <th>Elimination share</th>
          </tr>
        </thead>
        <tbody>
          {''.join(table_rows)}
        </tbody>
      </table>
    </section>
  </main>
</body>
</html>"""
    return html


def save_comparison_dashboard(payload: dict[str, Any]) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    COMPARISON_OUTPUT_FILE.write_text(render_comparison_html(payload), encoding="utf-8")


def save_daily_kpi_summary(records: list[dict[str, Any]]) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    DAILY_KPI_FILE.write_text(
        json.dumps(build_daily_kpi_summary(records), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def main(explicit_path: str | None = None) -> None:
    print("Startujem tvorbu dashboardu...")
    try:
        excel_path = resolve_excel_input_path(explicit_path)
    except FileNotFoundError as error:
        print(f"Chyba: {error}")
        print(f"Upload priecinok: {INPUT_DIR}")
        sys.exit(1)

    print(f"Nacitavam subor: {excel_path.name}")
    payload = build_payload_from_excel(excel_path)
    save_dashboard(payload)
    save_comparison_dashboard(payload)
    save_daily_kpi_summary(payload["records"])

    print("Hotovo.")
    print(f"Dashboard je ulozeny tu: {OUTPUT_FILE}")
    print(f"Porovnávací dashboard je ulozený tu: {COMPARISON_OUTPUT_FILE}")


if __name__ == "__main__":
    if is_streamlit_runtime():
        render_streamlit_dashboard()
    else:
        parser = argparse.ArgumentParser(description="Build packaging dashboard from Excel input.")
        parser.add_argument(
            "--input",
            dest="input_path",
            default="",
            help="Optional path to the Excel file. Defaults to the newest Excel in input/.",
        )
        args = parser.parse_args()
        main(args.input_path or None)
